from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
import psycopg
from psycopg.rows import dict_row
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from functools import wraps
import hashlib
import secrets
import re
import os
import logging
from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv
from collections import defaultdict, Counter
from time import time

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32))

if not os.path.exists('logs'):
    os.makedirs('logs')

file_handler = RotatingFileHandler('logs/app.log', maxBytes=10485760, backupCount=10)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
))
file_handler.setLevel(logging.INFO)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)
app.logger.info('Application startup')

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=False,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=int(os.environ.get('SESSION_LIFETIME', 86400))
)

DB_CONFIG = {
    'dbname': "camers",
    'user':  "camers",
    'password': "qwerty",
    'host': "web24.urbus.ru",
    'port': "5432",
}

ALLOWED_TABLES = {
    'cam_registrators', 'cam_camers', 'cam_camera_report',
    'cam_users', 'cam_action_log'
}

rate_limits = defaultdict(list)
active_sessions = {}

def rate_limit(max_requests=100, time_window=60):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            now = time()
            user_id = session.get('user_id', request.remote_addr)
            user_requests = rate_limits[user_id]
            while user_requests and user_requests[0] < now - time_window:
                user_requests.pop(0)
            if len(user_requests) >= max_requests:
                app.logger.warning(f'Rate limit exceeded for user {user_id}')
                return jsonify({'error': 'Too many requests. Please try again later.'}), 429
            user_requests.append(now)
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def validate_table_name(table_name):
    return table_name in ALLOWED_TABLES

def validate_column_name(column_name):
    return bool(re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', column_name))

def validate_record_data(table_name, data):
    errors = []
    if table_name == 'cam_camers':
        if 'port' in data:
            try:
                port = int(data['port'])
                if port < 1 or port > 32:
                    errors.append('Номер порта должен быть от 1 до 32')
            except (ValueError, TypeError):
                errors.append('Номер порта должен быть числом')
    elif table_name == 'cam_registrators':
        if 'ap' in data:
            try:
                ap = int(data['ap'])
                if ap < 1 or ap > 999:
                    errors.append('Номер АП должен быть от 1 до 999')
            except (ValueError, TypeError):
                errors.append('Номер АП должен быть числом')
        if 'count_ports' in data:
            try:
                ports = int(data['count_ports'])
                if ports not in [4, 8, 16, 32]:
                    errors.append('Количество портов должно быть 4, 8, 16 или 32')
            except (ValueError, TypeError):
                errors.append('Количество портов должно быть числом')
        if 'ip' in data and data['ip']:
            ip_pattern = re.compile(r'^(\d{1,3}\.){3}\d{1,3}$|^localhost$|^127\.0\.0\.1$')
            if not ip_pattern.match(data['ip']):
                errors.append('Неверный формат IP-адреса')
    elif table_name == 'cam_camera_report':
        if 'condition' in data:
            valid_conditions = ['Исправна', 'Частично не исправна', 'Неисправна', 'Отключена', 'Проба']
            if data['condition'] not in valid_conditions:
                errors.append('Неверное значение состояния')
    return errors

def get_db_connection():
    try:
        conn = psycopg.connect(**DB_CONFIG)
        conn.row_factory = dict_row
        return conn
    except Exception as e:
        app.logger.error(f'Database connection error: {e}')
        raise

def hash_password(password):
    if not password:
        return ''
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, hashed):
    if not password or not hashed:
        return password == hashed
    if hash_password(password) == hashed:
        return True
    if password == hashed:
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("UPDATE cam_users SET password = %s WHERE password = %s",
                       (hash_password(password), password))
            conn.commit()
            cur.close()
            conn.close()
            app.logger.info('Migrated plain text password to hash')
        except Exception as e:
            app.logger.error(f'Password migration error: {e}')
        return True
    return False

def format_datetime_for_display(dt_value):
    if not dt_value:
        return {'date': '', 'time': ''}
    if hasattr(dt_value, 'strftime'):
        return {
            'date': dt_value.strftime('%d.%m.%Y'),
            'time': dt_value.strftime('%H:%M:%S')
        }
    if isinstance(dt_value, str):
        try:
            dt = datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
            return {
                'date': dt.strftime('%d.%m.%Y'),
                'time': dt.strftime('%H:%M:%S')
            }
        except:
            return {'date': dt_value, 'time': ''}
    return {'date': str(dt_value), 'time': ''}

def log_action(user_id, action, table_name=None, record_id=None, field_name=None, old_value=None, new_value=None):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        if field_name == 'password':
            old_value = '***'
            new_value = '***'
        if old_value is not None and not isinstance(old_value, str):
            old_value = str(old_value)
        if new_value is not None and not isinstance(new_value, str):
            new_value = str(new_value)
        cur.execute("""
            INSERT INTO cam_action_log (time_action, user_id, action, table_name, record_id, field_name, old_value, new_value)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (datetime.now(), user_id, action, table_name, record_id, field_name, old_value, new_value))
        conn.commit()
        cur.close()
        conn.close()
        app.logger.info(f'Action logged: user_id={user_id} - {action}')
    except Exception as e:
        app.logger.error(f'Error logging action: {e}')

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'username' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        username = session['username']
        if username not in active_sessions or active_sessions[username] != session.get('session_token'):
            session.clear()
            return jsonify({'error': 'Session expired or logged in elsewhere'}), 401
        return f(*args, **kwargs)
    return wrapper

def can_view_table(table_name, role):
    if role == 'admin':
        return True
    if role == 'editor':
        return table_name not in ['cam_users', 'cam_action_log']
    if role == 'user':
        return table_name in ['cam_registrators', 'cam_camers', 'cam_camera_report']
    return False

def can_edit_table(table_name, role):
    if role == 'admin':
        return True
    if role == 'editor':
        return table_name == 'cam_camera_report'
    return False

@app.route('/login', methods=['GET', 'POST'])
@rate_limit(max_requests=10, time_window=300)
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("SELECT id, username, role, password FROM cam_users WHERE username = %s", (username,))
            user = cur.fetchone()
            cur.close()
            conn.close()
            if user and verify_password(password, user['password']):
                session.permanent = True
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['role'] = user['role']
                token = secrets.token_urlsafe(16)
                active_sessions[username] = token
                session['session_token'] = token
                log_action(session['user_id'], "Успешный вход в систему")
                app.logger.info(f'Successful login: {username}')
                return redirect(url_for('index'))
            else:
                log_action(None, "Неудачная попытка входа")
                app.logger.warning(f'Failed login attempt: {username}')
                return render_template('login.html', error="Неверное имя пользователя или пароль")
        except Exception as e:
            app.logger.error(f'Login error: {e}')
            return render_template('login.html', error="Ошибка подключения к базе данных")
    return render_template('login.html')

@app.route('/')
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', session=session)

@app.route('/logout')
def logout():
    username = session.get('username')
    user_id = session.get('user_id')
    if username and username in active_sessions:
        del active_sessions[username]
    log_action(user_id, "Выход из системы")
    app.logger.info(f'Logout: {username}')
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/data/<table_name>')
@login_required
@rate_limit(max_requests=100)
def get_table_data(table_name):
    if not validate_table_name(table_name):
        return jsonify({'error': 'Invalid table name'}), 400
    role = session.get('role', 'user')
    if not can_view_table(table_name, role):
        return jsonify({'error': 'Access denied'}), 403
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        if table_name == 'cam_action_log':
            cur.execute("""
                SELECT al.id, al.time_action, al.action, al.table_name, al.record_id,
                       al.field_name, al.old_value, al.new_value, u.username AS user
                FROM cam_action_log al
                LEFT JOIN cam_users u ON al.user_id = u.id
                ORDER BY al.id
            """)
        else:
            cur.execute(f'SELECT *, version FROM camers.{table_name} ORDER BY id')

        data = cur.fetchall()

        if table_name != 'cam_action_log':
            cur.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = %s ORDER BY ordinal_position
            """, (table_name,))
            columns = [col['column_name'] for col in cur.fetchall()]
        else:
            columns = ['action_date', 'action_time', 'user', 'action', 'table_name', 'record_id', 'field_name', 'old_value', 'new_value']

        cur.close()
        conn.close()

        formatted_data = []
        for row in data:
            row_dict = dict(row)
            if table_name == 'cam_action_log' and 'time_action' in row_dict:
                formatted_dt = format_datetime_for_display(row_dict['time_action'])
                row_dict['action_date'] = formatted_dt['date']
                row_dict['action_time'] = formatted_dt['time']
                del row_dict['time_action']
            if table_name == 'cam_users' and 'password' in row_dict:
                row_dict['password'] = '••••••'
            formatted_data.append(row_dict)

        return jsonify({'data': formatted_data, 'columns': columns})
    except Exception as e:
        app.logger.error(f'Error getting table data: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/<table_name>/<int:id>')
@login_required
@rate_limit(max_requests=100)
def get_record(table_name, id):
    if not validate_table_name(table_name):
        return jsonify({'error': 'Invalid table name'}), 400
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f'SELECT *, version FROM camers.{table_name} WHERE id = %s', (id,))
        record = cur.fetchone()
        cur.close()
        conn.close()
        if record:
            if table_name == 'cam_users' and 'password' in record:
                record['password'] = '••••••'
            return jsonify(record)
        return jsonify({'error': 'Record not found'}), 404
    except Exception as e:
        app.logger.error(f'Error getting record: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/structure/<table_name>')
@login_required
#@rate_limit(max_requests=300, time_window=60)
def get_table_structure(table_name):
    if not validate_table_name(table_name):
        return jsonify({'error': 'Invalid table name'}), 400
    role = session.get('role', 'user')
    if not can_edit_table(table_name, role):
        return jsonify({'error': 'Access denied'}), 403
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT column_name, is_nullable, data_type
            FROM information_schema.columns
            WHERE table_name = %s
            ORDER BY ordinal_position
        """, (table_name,))
        columns = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({'columns': [col['column_name'] + (' (NOT NULL)' if col['is_nullable'] == 'NO' else '') for col in columns]})
    except Exception as e:
        app.logger.error(f'Error getting table structure: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/<table_name>', methods=['POST'])
@login_required
@rate_limit(max_requests=50)
def add_record(table_name):
    if not validate_table_name(table_name):
        return jsonify({'error': 'Invalid table name'}), 400

    try:
        data = request.json
        if 'id' in data:
            del data['id']

        validation_errors = validate_record_data(table_name, data)
        if validation_errors:
            return jsonify({'success': False, 'error': ', '.join(validation_errors)}), 400

        if table_name == 'cam_users' and 'password' in data:
            data['password'] = hash_password(data['password'])

        for col in data.keys():
            if not validate_column_name(col):
                return jsonify({'success': False, 'error': 'Invalid column name'}), 400

        columns = list(data.keys())
        values = list(data.values())
        placeholders = ','.join(['%s'] * len(columns))
        columns_str = ','.join(columns)

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f"""
            INSERT INTO camers.{table_name} ({columns_str})
            VALUES ({placeholders}) RETURNING id
        """, values)
        new_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()

        log_action(session['user_id'], "Добавление записи", table_name, new_id)
        app.logger.info(f'Record added: {table_name}:{new_id} by user_id={session["user_id"]}')
        return jsonify({'success': True, 'id': new_id})

    except Exception as e:
        app.logger.error(f'Error adding record: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/data/<table_name>/<int:id>', methods=['PUT'])
@login_required
@rate_limit(max_requests=50)
def update_record(table_name, id):
    if not validate_table_name(table_name):
        return jsonify({'error': 'Invalid table name'}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f'SELECT * FROM camers.{table_name} WHERE id = %s', (id,))
        old_record = cur.fetchone()
        if not old_record:
            cur.close()
            conn.close()
            return jsonify({'error': 'Record not found'}), 404

        data = request.json
        if 'id' in data:
            del data['id']

        client_version = data.pop('version', None)
        current_version = old_record.get('version', 0)

        if client_version is not None and client_version != current_version:
            cur.close()
            conn.close()
            return jsonify({
                'success': False,
                'error': 'Запись была изменена другим пользователем. Обновите страницу и повторите попытку.'
            }), 409

        role = session.get('role')
        if role == 'editor' and table_name == 'cam_camera_report':
            allowed_fields = {'comment'}
            data = {k: v for k, v in data.items() if k in allowed_fields}
            if not data:
                cur.close()
                conn.close()
                return jsonify({'success': False, 'error': 'Редактор может изменять только поле "Примечание"'}), 403

        validation_errors = validate_record_data(table_name, data)
        if validation_errors:
            return jsonify({'success': False, 'error': ', '.join(validation_errors)}), 400

        if table_name == 'cam_users' and 'password' in data:
            if data['password'] == '••••••' or not data['password']:
                del data['password']
            else:
                data['password'] = hash_password(data['password'])

        for col in data.keys():
            if not validate_column_name(col):
                return jsonify({'success': False, 'error': 'Invalid column name'}), 400

        if not data:
            cur.close()
            conn.close()
            return jsonify({'error': 'No data to update'}), 400

        set_clause = ','.join([f"{key}=%s" for key in data.keys()])
        values = list(data.values()) + [id]
        cur.execute(f"UPDATE camers.{table_name} SET {set_clause} WHERE id = %s", values)
        conn.commit()

        user_id = session['user_id']
        for key, new_value in data.items():
            old_value = old_record.get(key)
            old_val_str = '' if old_value is None else str(old_value)
            new_val_str = '' if new_value is None else str(new_value)
            if old_val_str != new_val_str:
                log_action(user_id, "Изменение поля", table_name, id, key, old_val_str, new_val_str)

        cur.close()
        conn.close()
        app.logger.info(f'Record updated: {table_name}:{id} by user_id={user_id}')
        return jsonify({'success': True})

    except Exception as e:
        app.logger.error(f'Error updating record: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/data/<table_name>/<int:id>', methods=['DELETE'])
@login_required
@rate_limit(max_requests=30)
def delete_record(table_name, id):
    if not validate_table_name(table_name):
        return jsonify({'error': 'Invalid table name'}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        log_action(session['user_id'], "Удаление записи", table_name, id)
        cur.execute(f'DELETE FROM camers.{table_name} WHERE id = %s', (id,))
        conn.commit()
        cur.close()
        conn.close()
        app.logger.info(f'Record deleted: {table_name}:{id} by user_id={session["user_id"]}')
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f'Error deleting record: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/camers/cameras')
@login_required
@rate_limit(max_requests=100)
def get_camers_cameras():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, port, location, idreg FROM camers.cam_camers ORDER BY id')
        cameras = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify(cameras)
    except Exception as e:
        app.logger.error(f'Error getting cameras: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/camers/registrators')
@login_required
@rate_limit(max_requests=100)
def get_camers_registrators():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT id, ap, id_reg_on_ap FROM camers.cam_registrators ORDER BY id')
        registrators = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify(registrators)
    except Exception as e:
        app.logger.error(f'Error getting registrators: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/export_excel', methods=['POST'])
@login_required
def export_excel():
    data = request.json
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по камерам"

    headers = ['АП', 'Регистратор', 'Камера', 'Тип', 'Расположение', 'Расширение', 'Состояние', 'Тип поломки', 'Дата записи']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row in data:
        ws.append([
            row.get('АП', ''),
            row.get('Регистратор', ''),
            row.get('Камера', ''),
            row.get('Тип', ''),
            row.get('Расположение', ''),
            row.get('Расширение', ''),
            row.get('Состояние', ''),
            row.get('Тип поломки', ''),
            row.get('Дата записи', '')
        ])

    column_widths = [8, 12, 8, 12, 25, 12, 15, 20, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    app.logger.info(f'Excel export: camera_report by user_id={session["user_id"]}')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'camera_report_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    )

@app.route('/export_action_log_excel', methods=['POST'])
@login_required
def export_action_log_excel():
    data = request.json
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал действий"

    headers = ['Дата', 'Время', 'Пользователь', 'Действие', 'Таблица', 'ID записи', 'Поле', 'Было', 'Стало']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row in data:
        ws.append([
            row.get('Дата', ''),
            row.get('Время', ''),
            row.get('Пользователь', ''),
            row.get('Действие', ''),
            row.get('Таблица', ''),
            row.get('ID записи', ''),
            row.get('Поле', ''),
            row.get('Было', ''),
            row.get('Стало', '')
        ])

    column_widths = [12, 10, 15, 20, 15, 10, 15, 30, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    app.logger.info(f'Excel export: action_log by user_id={session["user_id"]}')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'action_log_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    )

# ========== АНАЛИТИКА (ОКОНЧАТЕЛЬНАЯ ВЕРСИЯ) ==========
@app.route('/api/analytics', methods=['GET'])
@login_required
@rate_limit(max_requests=50)
def get_analytics():
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    ap_ids = request.args.getlist('ap_ids')
    registrator_ids = request.args.getlist('registrator_ids')

    if not date_from_str or not date_to_str:
        return jsonify({'error': 'date_from and date_to required'}), 400

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    if date_from > date_to:
        return jsonify({'error': 'date_from must be <= date_to'}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    # Определяем тип фильтра
    use_ap_filter = False
    use_reg_filter = False
    filter_values = []
    if registrator_ids:
        use_reg_filter = True
        filter_values = [int(x) for x in registrator_ids]
    elif ap_ids:
        use_ap_filter = True
        filter_values = [int(x) for x in ap_ids]

    # --- KPI ---
    if use_reg_filter:
        cur.execute("""
            SELECT COUNT(DISTINCT c.id) as total_cameras,
                   COUNT(DISTINCT r.id) as total_registrators,
                   COUNT(DISTINCT r.ap) as total_aps
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
            WHERE r.id = ANY(%s::int[])
        """, (filter_values,))
    elif use_ap_filter:
        cur.execute("""
            SELECT COUNT(DISTINCT c.id) as total_cameras,
                   COUNT(DISTINCT r.id) as total_registrators,
                   COUNT(DISTINCT r.ap) as total_aps
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
            WHERE r.ap = ANY(%s::int[])
        """, (filter_values,))
    else:
        cur.execute("""
            SELECT COUNT(DISTINCT c.id) as total_cameras,
                   COUNT(DISTINCT r.id) as total_registrators,
                   COUNT(DISTINCT r.ap) as total_aps
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
        """)
    kpi_row = cur.fetchone()
    total_cameras = kpi_row['total_cameras'] or 0
    total_registrators = kpi_row['total_registrators'] or 0
    total_aps = kpi_row['total_aps'] or 0

    # % отчётов за последние 7 дней
    week_ago = date_to - timedelta(days=7)
    if use_reg_filter:
        cur.execute("""
            SELECT COUNT(DISTINCT c.id) as cnt
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
            WHERE EXISTS (
                SELECT 1 FROM cam_camera_report rep
                WHERE rep.id_cam = c.id AND rep.recording_date BETWEEN %s AND %s
            )
            AND r.id = ANY(%s::int[])
        """, (week_ago, date_to, filter_values))
    elif use_ap_filter:
        cur.execute("""
            SELECT COUNT(DISTINCT c.id) as cnt
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
            WHERE EXISTS (
                SELECT 1 FROM cam_camera_report rep
                WHERE rep.id_cam = c.id AND rep.recording_date BETWEEN %s AND %s
            )
            AND r.ap = ANY(%s::int[])
        """, (week_ago, date_to, filter_values))
    else:
        cur.execute("""
            SELECT COUNT(DISTINCT c.id) as cnt
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
            WHERE EXISTS (
                SELECT 1 FROM cam_camera_report rep
                WHERE rep.id_cam = c.id AND rep.recording_date BETWEEN %s AND %s
            )
        """, (week_ago, date_to))
    recent_count = cur.fetchone()['cnt'] or 0
    report_coverage_7d = round(recent_count / total_cameras * 100, 1) if total_cameras > 0 else 0

    # % исправных на date_to
    if use_reg_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, rep.condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.id = ANY(%s::int[])
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT COUNT(*) as total,
                   SUM(CASE WHEN condition = 'Исправна' THEN 1 ELSE 0 END) as healthy
            FROM last_reports
        """, (date_to, filter_values))
    elif use_ap_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, rep.condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.ap = ANY(%s::int[])
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT COUNT(*) as total,
                   SUM(CASE WHEN condition = 'Исправна' THEN 1 ELSE 0 END) as healthy
            FROM last_reports
        """, (date_to, filter_values))
    else:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, rep.condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT COUNT(*) as total,
                   SUM(CASE WHEN condition = 'Исправна' THEN 1 ELSE 0 END) as healthy
            FROM last_reports
        """, (date_to,))
    status_row = cur.fetchone()
    healthy_count = status_row['healthy'] or 0
    healthy_percent = round(healthy_count / total_cameras * 100, 1) if total_cameras > 0 else 0

    kpi = {
        'total_cameras': total_cameras,
        'total_registrators': total_registrators,
        'total_aps': total_aps,
        'report_coverage_7d': report_coverage_7d,
        'healthy_percent': healthy_percent
    }

    # --- Распределение состояний на date_to (с полным списком камер) ---
    if use_reg_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, COALESCE(rep.condition, 'Нет данных') as condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.id = ANY(%s::int[])
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT condition, array_agg(cam_id) as cam_ids, COUNT(*) as count
            FROM last_reports
            GROUP BY condition
        """, (date_to, filter_values))
    elif use_ap_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, COALESCE(rep.condition, 'Нет данных') as condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.ap = ANY(%s::int[])
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT condition, array_agg(cam_id) as cam_ids, COUNT(*) as count
            FROM last_reports
            GROUP BY condition
        """, (date_to, filter_values))
    else:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, COALESCE(rep.condition, 'Нет данных') as condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT condition, array_agg(cam_id) as cam_ids, COUNT(*) as count
            FROM last_reports
            GROUP BY condition
        """, (date_to,))
    status_rows = cur.fetchall()
    status_distribution = {}
    for row in status_rows:
        condition = row['condition']
        status_distribution[condition] = {
            'count': row['count'],
            'cam_ids': row['cam_ids'] if row['cam_ids'] else []
        }

    # --- Динамика состояний по дням ---
    if use_reg_filter:
        cur.execute("""
            SELECT DISTINCT c.id as cam_id
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
            WHERE r.id = ANY(%s::int[])
        """, (filter_values,))
    elif use_ap_filter:
        cur.execute("""
            SELECT DISTINCT c.id as cam_id
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
            WHERE r.ap = ANY(%s::int[])
        """, (filter_values,))
    else:
        cur.execute("""
            SELECT DISTINCT c.id as cam_id
            FROM cam_camers c
            JOIN cam_registrators r ON c.idreg = r.id
        """)
    camera_ids = [row['cam_id'] for row in cur.fetchall()]

    if not camera_ids:
        cur.close()
        conn.close()
        return jsonify({
            'kpi': kpi,
            'status_distribution': status_distribution,
            'daily_status': [],
            'top_breakdowns': [],
            'problem_registrators': [],
            'longest_breakdowns': []
        })

    cur.execute("""
        SELECT id_cam as cam_id, recording_date, condition
        FROM cam_camera_report
        WHERE id_cam = ANY(%s) AND recording_date <= %s
        ORDER BY id_cam, recording_date
    """, [camera_ids, date_to])
    all_reports = cur.fetchall()

    reports_by_cam = {}
    for rep in all_reports:
        cam_id = rep['cam_id']
        if cam_id not in reports_by_cam:
            reports_by_cam[cam_id] = []
        reports_by_cam[cam_id].append(rep)

    date_list = []
    current = date_from
    while current <= date_to:
        date_list.append(current.isoformat())
        current += timedelta(days=1)

    all_conditions = ['Исправна', 'Частично не исправна', 'Неисправна', 'Отключена', 'Проба']
    daily_counts = {date: {cond: 0 for cond in all_conditions} for date in date_list}

    for cam_id, reports in reports_by_cam.items():
        reports.sort(key=lambda x: x['recording_date'])
        last_cond = 'Нет данных'
        rep_idx = 0
        for date_str in date_list:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            while rep_idx < len(reports) and reports[rep_idx]['recording_date'] <= date_obj:
                last_cond = reports[rep_idx]['condition']
                rep_idx += 1
            if last_cond != 'Нет данных':
                daily_counts[date_str][last_cond] += 1

    daily_status = []
    for date in date_list:
        entry = {'date': date}
        for cond in all_conditions:
            entry[cond] = daily_counts[date][cond]
        daily_status.append(entry)

    # --- Топ поломок ---
    if use_reg_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, rep.breakdown, rep.condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.id = ANY(%s::int[])
                  AND rep.condition IN ('Частично не исправна', 'Неисправна')
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT breakdown
            FROM last_reports
            WHERE breakdown IS NOT NULL AND breakdown != ''
        """, (date_to, filter_values))
    elif use_ap_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, rep.breakdown, rep.condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.ap = ANY(%s::int[])
                  AND rep.condition IN ('Частично не исправна', 'Неисправна')
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT breakdown
            FROM last_reports
            WHERE breakdown IS NOT NULL AND breakdown != ''
        """, (date_to, filter_values))
    else:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, rep.breakdown, rep.condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE rep.condition IN ('Частично не исправна', 'Неисправна')
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT breakdown
            FROM last_reports
            WHERE breakdown IS NOT NULL AND breakdown != ''
        """, (date_to,))
    breakdowns = []
    for row in cur.fetchall():
        parts = row['breakdown'].split(',')
        for part in parts:
            part = part.strip()
            if part:
                breakdowns.append(part)
    breakdown_counter = Counter(breakdowns)
    top_breakdowns = [{'breakdown': k, 'count': v} for k, v in breakdown_counter.most_common(10)]

    # --- Проблемные регистраторы ---
    if use_reg_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, c.idreg as reg_id, COALESCE(rep.condition, 'Нет данных') as condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.id = ANY(%s::int[])
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT r.id as reg_id,
                   CONCAT('АП', r.ap, '_', r.id_reg_on_ap) as reg_name,
                   COUNT(lr.cam_id) as total_cameras,
                   SUM(CASE WHEN lr.condition IN ('Частично не исправна', 'Неисправна') THEN 1 ELSE 0 END) as unhealthy
            FROM cam_registrators r
            LEFT JOIN last_reports lr ON lr.reg_id = r.id
            GROUP BY r.id, r.ap, r.id_reg_on_ap
            ORDER BY unhealthy DESC
        """, (date_to, filter_values))
    elif use_ap_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, c.idreg as reg_id, COALESCE(rep.condition, 'Нет данных') as condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.ap = ANY(%s::int[])
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT r.id as reg_id,
                   CONCAT('АП', r.ap, '_', r.id_reg_on_ap) as reg_name,
                   COUNT(lr.cam_id) as total_cameras,
                   SUM(CASE WHEN lr.condition IN ('Частично не исправна', 'Неисправна') THEN 1 ELSE 0 END) as unhealthy
            FROM cam_registrators r
            LEFT JOIN last_reports lr ON lr.reg_id = r.id
            GROUP BY r.id, r.ap, r.id_reg_on_ap
            ORDER BY unhealthy DESC
        """, (date_to, filter_values))
    else:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, c.idreg as reg_id, COALESCE(rep.condition, 'Нет данных') as condition
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                ORDER BY c.id, rep.recording_date DESC
            )
            SELECT r.id as reg_id,
                   CONCAT('АП', r.ap, '_', r.id_reg_on_ap) as reg_name,
                   COUNT(lr.cam_id) as total_cameras,
                   SUM(CASE WHEN lr.condition IN ('Частично не исправна', 'Неисправна') THEN 1 ELSE 0 END) as unhealthy
            FROM cam_registrators r
            LEFT JOIN last_reports lr ON lr.reg_id = r.id
            GROUP BY r.id, r.ap, r.id_reg_on_ap
            ORDER BY unhealthy DESC
        """, (date_to,))
    problem_registrators = []
    for row in cur.fetchall():
        total = row['total_cameras'] or 0
        unhealthy = row['unhealthy'] or 0
        percent = round(unhealthy / total * 100, 1) if total > 0 else 0
        problem_registrators.append({
            'registrator_name': row['reg_name'],
            'total_cameras': total,
            'unhealthy': unhealthy,
            'unhealthy_percent': percent
        })

    # --- Камеры с самой длительной непрерывной поломкой ---
    if use_reg_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, c.idreg, c.port, c.location,
                       rep.condition, rep.breakdown, rep.recording_date as last_report_date
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.id = ANY(%s::int[])
                ORDER BY c.id, rep.recording_date DESC
            ),
            first_broken AS (
                SELECT lr.cam_id, MIN(rep.recording_date) as start_date
                FROM cam_camera_report rep
                JOIN last_reports lr ON lr.cam_id = rep.id_cam
                WHERE rep.condition IN ('Частично не исправна', 'Неисправна')
                  AND rep.recording_date <= %s
                GROUP BY lr.cam_id
            )
            SELECT lr.cam_id, lr.idreg, lr.port, lr.location, lr.condition, lr.breakdown,
                   fb.start_date, (%s - fb.start_date) as days,
                   CONCAT('АП', r.ap, '_', r.id_reg_on_ap) as reg_name
            FROM last_reports lr
            JOIN first_broken fb ON lr.cam_id = fb.cam_id
            JOIN cam_registrators r ON lr.idreg = r.id
            WHERE lr.condition IN ('Частично не исправна', 'Неисправна')
              AND NOT EXISTS (
                  SELECT 1 FROM cam_camera_report rep2
                  WHERE rep2.id_cam = lr.cam_id
                    AND rep2.recording_date > fb.start_date AND rep2.recording_date <= %s
                    AND rep2.condition = 'Исправна'
              )
            ORDER BY days DESC
            LIMIT 10
        """, (date_to, filter_values, date_to, date_to, date_to))
    elif use_ap_filter:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, c.idreg, c.port, c.location,
                       rep.condition, rep.breakdown, rep.recording_date as last_report_date
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                WHERE r.ap = ANY(%s::int[])
                ORDER BY c.id, rep.recording_date DESC
            ),
            first_broken AS (
                SELECT lr.cam_id, MIN(rep.recording_date) as start_date
                FROM cam_camera_report rep
                JOIN last_reports lr ON lr.cam_id = rep.id_cam
                WHERE rep.condition IN ('Частично не исправна', 'Неисправна')
                  AND rep.recording_date <= %s
                GROUP BY lr.cam_id
            )
            SELECT lr.cam_id, lr.idreg, lr.port, lr.location, lr.condition, lr.breakdown,
                   fb.start_date, (%s - fb.start_date) as days,
                   CONCAT('АП', r.ap, '_', r.id_reg_on_ap) as reg_name
            FROM last_reports lr
            JOIN first_broken fb ON lr.cam_id = fb.cam_id
            JOIN cam_registrators r ON lr.idreg = r.id
            WHERE lr.condition IN ('Частично не исправна', 'Неисправна')
              AND NOT EXISTS (
                  SELECT 1 FROM cam_camera_report rep2
                  WHERE rep2.id_cam = lr.cam_id
                    AND rep2.recording_date > fb.start_date AND rep2.recording_date <= %s
                    AND rep2.condition = 'Исправна'
              )
            ORDER BY days DESC
            LIMIT 10
        """, (date_to, filter_values, date_to, date_to, date_to))
    else:
        cur.execute("""
            WITH last_reports AS (
                SELECT DISTINCT ON (c.id) c.id as cam_id, c.idreg, c.port, c.location,
                       rep.condition, rep.breakdown, rep.recording_date as last_report_date
                FROM cam_camers c
                JOIN cam_registrators r ON c.idreg = r.id
                LEFT JOIN cam_camera_report rep ON rep.id_cam = c.id AND rep.recording_date <= %s
                ORDER BY c.id, rep.recording_date DESC
            ),
            first_broken AS (
                SELECT lr.cam_id, MIN(rep.recording_date) as start_date
                FROM cam_camera_report rep
                JOIN last_reports lr ON lr.cam_id = rep.id_cam
                WHERE rep.condition IN ('Частично не исправна', 'Неисправна')
                  AND rep.recording_date <= %s
                GROUP BY lr.cam_id
            )
            SELECT lr.cam_id, lr.idreg, lr.port, lr.location, lr.condition, lr.breakdown,
                   fb.start_date, (%s - fb.start_date) as days,
                   CONCAT('АП', r.ap, '_', r.id_reg_on_ap) as reg_name
            FROM last_reports lr
            JOIN first_broken fb ON lr.cam_id = fb.cam_id
            JOIN cam_registrators r ON lr.idreg = r.id
            WHERE lr.condition IN ('Частично не исправна', 'Неисправна')
              AND NOT EXISTS (
                  SELECT 1 FROM cam_camera_report rep2
                  WHERE rep2.id_cam = lr.cam_id
                    AND rep2.recording_date > fb.start_date AND rep2.recording_date <= %s
                    AND rep2.condition = 'Исправна'
              )
            ORDER BY days DESC
            LIMIT 10
        """, (date_to, date_to, date_to, date_to))
    longest_breakdowns = []
    for row in cur.fetchall():
        longest_breakdowns.append({
            'camera_name': f"{row['reg_name']}_{row['port']}",
            'location': row['location'] or '',
            'condition': row['condition'],
            'breakdown': row['breakdown'] or '',
            'days': row['days'],
            'start_date': row['start_date'].isoformat()
        })

    cur.close()
    conn.close()

    return jsonify({
        'kpi': kpi,
        'status_distribution': status_distribution,
        'daily_status': daily_status,
        'top_breakdowns': top_breakdowns,
        'problem_registrators': problem_registrators,
        'longest_breakdowns': longest_breakdowns
    })

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=8080)